import pandas as pd
from datetime import (date, datetime)
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from openpyxl.chart import (ScatterChart, Reference, Series)
from openpyxl.chart.axis import DateAxis
import numpy as np
import xlsxwriter

# DOUBLE UNDERSCORE 
class Data_Processing: 
    CSV = '.csv'
    XLSX = '.xlsx'
    
    def __init__(self, choice, config_file, input_csv, output_wb): 
        self.choice = choice
        self.config_file = config_file
        self.input_csv = input_csv 
        self.output_wb = output_wb
    
    # Create the dataframe that stores the raw data of the CSV file 
    def create_csv_dataframe(self, file, startLine): 
        if (startLine.dropna().empty): 
            startLine = 0
        else: 
            startLine = startLine.loc[0]
        
        df = pd.read_csv(file + '.csv', header = startLine, keep_default_na = False)
        return df
    def create_excel_dataframe(self, file, sheet): 
        df = pd.read_excel(file + '.xlsx', sheet_name = sheet)
        #df = file[sheet]
        return df    
    # Create the Excel workbook that stores the raw data in Excel 
    def create_raw_Excelbook(self, data_df):  
        wb = Workbook()
        ws = wb.active
        ws.title = 'Raw Data'

        for row in dataframe_to_rows(data_df, index = False, header = True):
            ws.append(row)
        wb.save(self.get_input_csv + '.xlsx')
        return wb

    def create_plotted_workbook(self): 
        wb = Workbook()
        ws = wb.active
        ws.title = 'Output Data'
        return wb
    
    # Converts a column letter to its corresponding integer.
    # https://www.geeksforgeeks.org/find-excel-column-number-column-title/
    def letter2int(self, letters):
        result = 0
        for x in letters: 
            x = x.upper()
            result *= 26
            result += ord(x) - ord('A') + 1
        return result 

    # Converts the location format of the input columns from letters to new_titles  
    def letter2title(self, letters, names):    
        index = self.letter2int(letters)
        title = names[index-1]
        return title

    def time_format(self, datetime_series): 
        start_datetime = datetime_series.loc[0]
        start_list = start_datetime.split()
        start_time = pd.to_timedelta(start_list[1])
        for cur_datetime in datetime_series: 

            # Split the datetime string into a list by a space delimiter 
            # and store the HH:MM:SS portion into a variable. 
            cur_datetime_list = cur_datetime.split() 

            # Store the time portion into cur_time and convert it to a timedelta object 
            cur_time = pd.to_timedelta(cur_datetime_list[1])

            # Find the difference between the current time and the start time. 
            # Convert the timedelta object into a string and split string into a list
            # by space delimiter.  
            difference= str(cur_time-start_time)
            difference_list = difference.split()

            # Store the time portion of the string into elapsed_time
            elapsed_time = difference_list[2]

            # Convert elapsed_time to a datetime object and store the result in the date column 
            elapsed_time = datetime.strptime(elapsed_time, "%H:%M:%S").time()
            datetime_series.replace(cur_datetime, elapsed_time, inplace = True)
        return datetime_series

    # Store the data of the input columns of the CSV into the desired output columns in Excel 
    # new_titles will be to create the new names of the columns
    # num_inputs will be used to locate the cells where we want to store our data 
    # title_inputs will be used to retrieve the column datas 
    def process_data(self, wb, df, config_df):

        new_titles = config_df['Title']
        title_inputs = config_df['Input Column Title']
        outputs = config_df['Output']
        ranges = config_df['Range']

        # ENCAPSULATE INTO A NEW FUNCTION `
        # Read in all the data 
        for j in range(new_titles.size): 
            # Append and bold the header of input column to the first row of its desired column location in Excel. 
            self.read_in_values(wb, j, df, new_titles, title_inputs, outputs, ranges)
        return wb
    
    # Read in the value of 1 column to the output file 
    def read_in_values(self, wb, j, df, new_titles, title_inputs, outputs, range): 
        ws = wb.active
        header = ws.cell(row=1, column = outputs.iloc[j]) 
        header.value = new_titles.iloc[j]
        header.font = Font(bold=True)
        col_index = title_inputs.iloc[j]
        
        ## Range
        max_size = df[col_index].size
        current_range = self.find_range(range.loc[j],max_size)
        
        start = current_range[0]
        end = current_range[1]

        # Indices: i helps to retrieve the contents of the current column 
        #          cellRow helps ensures that the contents are placed in the correct cell 
        i = start
        cellRow = 2 
        while (i <= end):  
            ws.cell(row = cellRow, column = outputs.loc[j]).value = df.loc[i,col_index]
            cellRow += 1
            i += 1

    # Determine the starting and ending point of the data to be read 
    # Range is calculated against the row indexes of the Excel worksheet. Thus, the first
    # cell in a column will be located at row 2  
    def find_range(self, current_range, total_size): 
        if (pd.isnull(current_range)): 
            return [0,total_size-1]
        else: 
            range_list = current_range.split(':')
            start = int(range_list[0])-2
            end = int(range_list[1])-2
            if (start < 0): 
                return [0, total_size-1]
            return [start,end]

        
    # Effectively determines whether or not a chart will be created. 
    def make_chart(self,axis):

        # Extract the row index (if any) of the value that will serve as our x-axis 
        x_axis = axis.loc[(axis == 'x') | (axis == 'X')]
        return x_axis

    def create_chart(self,wb, outputs_data_df, x_axis, y_axis, config_df_1, config_df_2): 

        ws = wb.active
        
        title_inputs = config_df_1['Input Column Title']
        outputs = config_df_1['Output']
        new_titles = config_df_1['Title']
        graph_title = config_df_2['Graph Title']

        # Assume number of rows will be same throughout dataframe 
        row_size = outputs_data_df[title_inputs.loc[0]].size
        
        cs = wb.create_chartsheet()
        chart = ScatterChart()

        # Store the index location of the x-axis value 
        x_axis_row= x_axis.index[0] 

        # Store the column number where the x_axis is located 
        x = Reference(ws, min_col=outputs.loc[x_axis_row], min_row = 2, max_row = row_size)
        
        # Plot as many y-axes as indicated in the configuration file 
    
        y_axis_rows = y_axis.index
        
        for row in y_axis_rows: 
            y = Reference(ws, min_col = outputs.loc[row], min_row = 2, max_row = row_size)
            s = Series(y,x,title=new_titles.loc[row])
            chart.append(s)
        
    
        chart.x_axis.title = new_titles.loc[x_axis_row]
        # situate x-axis below negative numbers 
        chart.x_axis.tickLblPos = "low"

        #chart.x_axis.tickLblSkip = 3
        self.chart_legend(chart, y_axis_rows, new_titles)
        self.chart_title(chart, new_titles, graph_title, x_axis_row, y_axis_rows)
        self.chart_scaling(chart, config_df_2['X Min'], config_df_2['X Max'], config_df_2['Y Min'], config_df_2['Y Max'])
        cs.add_chart(chart)


    def chart_scaling(self, chart, x_min, x_max, y_min, y_max): 
        if (not x_min.dropna().empty): 
            chart.x_axis.scaling.min = x_min.loc[0]
        if (not x_max.dropna().empty): 
            chart.x_axis.scaling.max = x_max.loc[0]
        if (not y_min.dropna().empty): 
            chart.y_axis.scaling.min = y_min.loc[0]
        if (not y_max.dropna().empty): 
            chart.y_axis.scaling.max = y_max.loc[0]
        
    # Determines the need for a chart legend
    #   If there is only 1 y-axis, title the y_axis and delete the legend  

    def chart_legend(self,chart, y_axis_rows, new_titles):
        if (len(y_axis_rows) == 1): 
            chart.y_axis.title = new_titles.loc[y_axis_rows[0]]
            chart.legend = None 
        return None 

    ### Default chart title: If there is no given chart title then chart title will be: 
        #   'All y-axis vs x-axis'
    def chart_title(self,chart, new_titles, graph_title, x_axis_row, y_axis_rows): 
        # Note: A column with 'NaNs' is not considered empty. 
        if (graph_title.dropna().empty): 
            title = ' '
            for i in range(y_axis_rows.size-1): 
                title += new_titles.loc[y_axis_rows[i]] + ", "
            title += new_titles.loc[y_axis_rows[y_axis_rows.size-1]] + " vs " + new_titles.loc[x_axis_row]
            chart.title = title
        else: 
            chart.title = graph_title.loc[0]

    def read_config_file(self): 
        read_file = pd.read_excel('LumenConfig.xlsx', sheet_name = 'Sheet1')
        return read_file

    # Convert the letter elements of inputs into integers and Strings and outputs into integers 
    # so we can later use them as indices. 
    def convert_columns(self, config_df, col_names):
        title_inputs = config_df['Input'].copy()    
        config_df['Input Column Title'] = title_inputs
        title_inputs = config_df['Input Column Title']
        num_inputs = config_df['Input']
        outputs = config_df['Output']
        for i in range(0, num_inputs.size): 
            num_inputs.replace(num_inputs.loc[i], self.letter2int(num_inputs.loc[i]), inplace = True)
            title_inputs.replace(title_inputs.loc[i], self.letter2title(title_inputs.loc[i], col_names), inplace = True)
            outputs.replace(outputs.loc[i], self.letter2int(outputs.loc[i]), inplace = True)
        return config_df
    
    def create_mapping_dataframe(self, raw_data_df, title_inputs):
        df = raw_data_df[[title_inputs.loc[0]]]
        for i in range(1, title_inputs.size): 
            new_df = raw_data_df[[title_inputs.loc[i]]]
            df = df.join(new_df)
        return df


    # getters and setters 
    
    @property
    def get_config_file(self): 
        return self.config_file
    @get_config_file.setter
    def set_config_file(self, config_file): 
        self.config_file = config_file

    @property
    def get_choice(self):
        return self.choice
    @get_choice.setter
    def set_choice(self, choice): 
        self.choice = choice

    @property
    def get_input_csv(self): 
        return self.input_csv
    @get_input_csv.setter
    def set_input_csv(self, input_csv): 
        self.input_csv = input_csv 
    
    ##### get_output_wb returns an object, not a String
    @property
    def get_output_wb(self): 
        return self.output_wb 
    @get_output_wb.setter
    def set_output_wb(self, output_wb): 
        self.output_wb = output_wb

    

